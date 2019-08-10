import boto3
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Border, Side

# Variables
awsAccID = '123456789012' # AWS account number
assumeRole = 'arn:' # ARN of role that you want to assume on account above
vpcId = 'vpc-' # VPC ID from account above
fileName = 'SecurityGroups - AWS ' + awsAccID + '.xlsx' # Name of .xlsx file
tokenSerial = '' # In case of MFA - Serial number of hardware token

# Assume role on specific account and get back ec2 client credentials
def login(awsAccID, prefix):
    token = input('Enter token code: ')
    clientSTS = boto3.client('sts')
    stsCreds = clientSTS.assume_role(
                RoleArn=assumeRole,
                RoleSessionName='SecurityGroupsExport',
                DurationSeconds=3600,
                SerialNumber=tokenSerial,
                TokenCode=str(token)
            )

    client = boto3.client(
            'ec2',
            aws_access_key_id=stsCreds['Credentials']['AccessKeyId'],
            aws_secret_access_key=stsCreds['Credentials']['SecretAccessKey'],
            aws_session_token=stsCreds['Credentials']['SessionToken'],
        )

    return client

# Visual modifications of xlsx to make it easier to read
def style(ws, value, row, column, color):
    ws.cell(row=row, column=column).value = value
    ws.cell(row=row, column=column).font = Font(bold=True,size=12)
    ws.cell(row=row, column=column).fill = PatternFill("solid", fgColor=color)
    ws.cell(row=row, column=column).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

# Export of ec2 list with some details
def ec2_export(awsAccID, prefix, vpcId, fileName):
    client = login(awsAccID, prefix)
    ec2list = client.describe_instances()

    try:
        wb = load_workbook(filename=fileName)
        ws = wb.create_sheet(title="Ec2 List")
    except:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ec2 List"

    i = 0
    for ec2 in ec2list["Reservations"]:
        if ec2['Instances'][0]['NetworkInterfaces'][0]['VpcId'] == vpcId:
            for tag in ec2['Instances'][0]["Tags"]:
                if tag["Key"] == "Name":
                    style(ws, 'Name', i + 2, 2, "000000FF")
                    ws.cell(row=2 + i, column=3).value = ec2['Instances'][0]["Tags"][ec2['Instances'][0]["Tags"].index(tag)]["Value"]
                    break

            style(ws, 'InstanceId', 3 + i, 2, "000000FF")
            ws.cell(row=3 + i, column=3).value = ec2['Instances'][0]["InstanceId"]

            style(ws, 'InstanceType', 4 + i, 2, "E59000")
            ws.cell(row=4 + i, column=3).value = ec2['Instances'][0]["InstanceType"]

            style(ws, 'PrivateIpAddress', 5 + i, 2, "E59000")
            ws.cell(row=5 + i, column=3).value = ec2['Instances'][0]["PrivateIpAddress"]

            style(ws, 'State', 6 + i, 2, "E59000")
            ws.cell(row=6 + i, column=3).value = ec2['Instances'][0]["State"]["Name"]

            style(ws, 'SecurityGroups:', 7 + i, 2, "E59000")
            for sg in ec2['Instances'][0]['SecurityGroups']:
                ws.cell(row=7 + i, column=3).value = sg["GroupName"]
                i = i + 1

            i = i + 7

        wb.save(fileName)

# main
ec2_export(awsAccID, prefix, vpcId, fileName)
